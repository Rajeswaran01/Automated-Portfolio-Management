# -*- coding: utf-8 -*-
"""
Created on Fri Feb 24 20:01:42 2023

@author: a-raj
"""

import os
import openpyxl


path="chemin d'accès"
Liste_titre=["zéro"]
for nom in os.listdir(path):
    sigle=nom[:nom.index(".")]
    if( sigle not in Liste_titre):
        Liste_titre.append(sigle)
        
Liste_titre.sort()
taille=len(Liste_titre)

wb = openpyxl.load_workbook('nom_du_fichier.xlsx')
feuille=wb["Feuille1"]

for i in range(1,taille):
    k=1
    for j in range(2012,2022):
        année=str(j)
        nom1= Liste_titre[i]+".PA_S1_"+année
        nom2= Liste_titre[i]+".PA_S2_"+année
        cellule=feuille.cell(k,i)
        if(nom1 in os.listdir(path)):
            cellule.value=Score(nom1)
        else:
            cellule.value=0
        k=k+1
        cellule=feuille.cell(k,i)
        if(nom2 in os.listdir(path)):
            cellule.value=Score(nom1)
        else:
            cellule.value=0
        k=k+1
wb.save("nom_du_fichier.xlsx")
ligne=10


def Actualisation( annee,  semestre,ligne,taille, Liste_titre):
    if(annee>2021 and (semestre =="S1" or semestre =="S2")):
       wb = openpyxl.load_workbook('nom_du_fichier.xlsx')
       feuille=wb["Feuille1"]
       ligne=ligne+1
       for i in range(1,taille):
           année=str(annee)
           nom1= Liste_titre[i]+".PA_"+ semestre+"_"+année
           nom2= Liste_titre[i]+".PA_"+ semestre + "_"+année
           cellule=feuille.cell(ligne,i)
           if(nom1 in os.listdir(path)):
               cellule.value=Score(nom1)
           else:
               cellule.value=0
           cellule=feuille.cell(k,i)
           if(nom2 in os.listdir(path)):
               cellule.value=Score(nom1)
           else:
               cellule.value=0
    wb.save("nom_du_fichier.xlsx")
        

           
           
       
       
       